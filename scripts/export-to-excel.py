"""
Exporteer alle ESG Quickscan-content naar een Excel-bestand dat Act Right
tekstueel kan reviewen en terugleveren voor implementatie.

Output: ESG-Quickscan-review.xlsx in de repo-root.

Tabbladen:
  1. Leeswijzer  - hoe je dit document gebruikt
  2. Vragen      - alle 20 vragen + antwoordopties
  3. Topics      - alle 98 ESG-onderwerpen (wetgeving/standaarden/frameworks)
  4. Rules       - alle 120+ if-then regels die topics koppelen aan antwoorden
  5. Themas      - de 8 overkoepelende thema's

Editeerbaar = groene achtergrond in kopregel.
Niet editeerbaar (keys/references) = grijze achtergrond.
"""

import re
import json5
import os
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO = Path(__file__).parent.parent
DATA = REPO / "src" / "data"
OUTPUT = REPO / "ESG-Quickscan-review.xlsx"

# --- Stijlen ---
HEADER_EDIT = PatternFill("solid", fgColor="C5D63D")    # Act Right lime - editable
HEADER_KEY  = PatternFill("solid", fgColor="DADCCD")    # grijs - keys/references
HEADER_INFO = PatternFill("solid", fgColor="384026")    # olive - intro/info
HEADER_FONT_LIGHT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FONT_DARK  = Font(bold=True, color="384026", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="DADCCD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def extract_array(ts_text: str, var_name: str) -> str:
    """Haal de array-literal uit `export const NAME: Type = [ ... ];`."""
    # find opening bracket
    m = re.search(rf'export const {var_name}[^=]*=\s*(\[)', ts_text)
    if not m:
        raise RuntimeError(f"Kon array '{var_name}' niet vinden")
    start = m.start(1)
    # balance brackets to find matching ]
    depth = 0
    i = start
    in_string = False
    string_char = None
    while i < len(ts_text):
        c = ts_text[i]
        if in_string:
            if c == '\\':
                i += 2
                continue
            if c == string_char:
                in_string = False
        else:
            if c in '"\'':
                in_string = True
                string_char = c
            elif c == '[':
                depth += 1
            elif c == ']':
                depth -= 1
                if depth == 0:
                    return ts_text[start:i + 1]
        i += 1
    raise RuntimeError(f"Geen balanced ] voor '{var_name}'")


def parse_ts_array(path: Path, var_name: str) -> list:
    """Lees TS-bestand, extraheer array, parse als JSON5."""
    text = path.read_text(encoding="utf-8")
    arr_text = extract_array(text, var_name)
    return json5.loads(arr_text)


# --- Data laden ---
print("Lezen van bron-bestanden...")
questions = parse_ts_array(DATA / "questions.ts", "questions")
topics = parse_ts_array(DATA / "topics.ts", "topics")
rules = parse_ts_array(DATA / "rules.ts", "rules")
themes = parse_ts_array(DATA / "themes.ts", "themes")
print(f"  Vragen: {len(questions)}")
print(f"  Topics: {len(topics)}")
print(f"  Rules:  {len(rules)}")
print(f"  Themas: {len(themes)}")

# --- Workbook opzetten ---
wb = Workbook()

# ========= Tabblad 1: Leeswijzer =========
ws = wb.active
ws.title = "Leeswijzer"
ws.column_dimensions["A"].width = 110
ws["A1"] = "ESG Quickscan Light — reviewbestand"
ws["A1"].font = Font(bold=True, size=18, color="384026")

intro = [
    "",
    "Dit bestand bevat alle content van de ESG Quickscan Light zoals die live staat",
    f"(bron: GitHub-main, versie t/m commit van {os.popen('git -C ' + str(REPO) + ' rev-parse --short HEAD').read().strip() or 'onbekend'}).",
    "",
    "Hoe te gebruiken",
    "1. Lees de tabbladen 'Vragen', 'Topics', 'Rules' en 'Themas' door.",
    "2. Pas teksten aan in cellen met een GROENE kopregel (deze zijn editeerbaar).",
    "3. Cellen met een GRIJZE kopregel zijn keys/references — LAAT ZE STAAN.",
    "   Bijvoorbeeld: topic-ID, rule-ID, activatedTopicIds, questionId, trigger-waardes.",
    "   Als je deze aanpast breekt de koppeling tussen rules en topics.",
    "4. Voor nieuwe topics of rules: voeg ze onderaan toe, ik maak daar nieuwe ID's voor.",
    "   Geef in kolom A een aanduiding zoals 'NIEUW' of 'TOEVOEGEN'.",
    "5. Voor topics die je wilt verwijderen: zet 'VERWIJDEREN' in kolom A.",
    "6. Stuur het bijgewerkte bestand terug, ik verwerk de wijzigingen in een nieuwe commit.",
    "",
    "Kleurcodering kopregels",
    "  GROEN (Act Right-lime)      → aanpasbare content",
    "  GEEL                        → Opmerking-kolom — typ hier je feedback per rij",
    "  GRIJS                       → key/reference — niet wijzigen",
    "",
    "Tabblad-overzicht",
    "  Vragen   — 20 vragen van de Quickscan met antwoordopties",
    "  Topics   — 98 ESG-onderwerpen (wetgeving/standaarden/frameworks)",
    "  Rules    — 120+ if-then regels die bepalen welke topics getriggerd worden",
    "  Themas   — 8 overkoepelende thema's",
    "",
    "Belangrijke velden om kritisch te reviewen",
    "  Topics.recommendedLabel  — 'Nu verplicht' / 'Sectorspecifiek' / 'Monitoren' / etc.",
    "                              (bepaalt hoe stellig de Quickscan het topic presenteert)",
    "  Topics.description       — de korte uitleg van het onderwerp (in PDF/rapport)",
    "  Topics.whenRelevant      — tweede zin van 'Waarom voor u relevant' in rapport",
    "  Rules.reportPhrasing     — eerste zin van 'Waarom voor u relevant' in rapport",
    "  Rules.preferredLabel     — label dat de rule vots; wordt nog door de engine gecapt",
    "",
    "Terugkoppeling",
    "  Stuur het bestand naar mij retour (chat/e-mail). Ik implementeer de wijzigingen,",
    "  test dat alles consistent blijft, en push naar main voor een nieuwe Lovable-rebuild.",
]
for i, line in enumerate(intro, start=2):
    ws.cell(row=i, column=1, value=line)
    if line in ("Hoe te gebruiken", "Kleurcodering kopregels", "Tabblad-overzicht",
                "Belangrijke velden om kritisch te reviewen", "Terugkoppeling"):
        ws.cell(row=i, column=1).font = Font(bold=True, size=13, color="384026")


# ========= Tabblad 2: Vragen =========
ws = wb.create_sheet("Vragen")
# Verzamel headers: Q-id is key; question+helper+options editable
columns = [
    ("id",          "KEY",  14, "Vraag-ID, gebruikt in rules"),
    ("question",    "EDIT", 70, "De vraagtekst die de klant ziet"),
    ("type",        "KEY",  12, "single = 1 antwoord, multi = meerdere"),
    ("required",    "KEY",  10, "true/false"),
    ("helper",      "EDIT", 70, "Toelichting onder de vraag (wordt aan klant getoond)"),
    ("options",     "EDIT", 90, "Antwoordopties — 1 per regel, formaat: value | label"),
    ("Opmerking",   "NOTE", 60, "Typ hier je wijzigingsvoorstel of opmerking voor Act Right / Claude"),
]
HEADER_NOTE = PatternFill("solid", fgColor="FDE68A")    # zachte gele opmerking-kolom

def apply_header(ws_local, col_idx, name, kind, width, descr):
    c = ws_local.cell(row=1, column=col_idx, value=name)
    if kind == "EDIT":
        c.fill = HEADER_EDIT
    elif kind == "NOTE":
        c.fill = HEADER_NOTE
    else:
        c.fill = HEADER_KEY
    c.font = HEADER_FONT_DARK
    c.alignment = WRAP
    c.border = BORDER
    ws_local.column_dimensions[get_column_letter(col_idx)].width = width
    from openpyxl.comments import Comment
    c.comment = Comment(descr, "Act Right review")

for col_idx, (name, kind, width, descr) in enumerate(columns, start=1):
    apply_header(ws, col_idx, name, kind, width, descr)

for row_idx, q in enumerate(questions, start=2):
    ws.cell(row=row_idx, column=1, value=q.get("id", ""))
    ws.cell(row=row_idx, column=2, value=q.get("question", ""))
    ws.cell(row=row_idx, column=3, value=q.get("type", ""))
    ws.cell(row=row_idx, column=4, value=str(q.get("required", False)).lower())
    ws.cell(row=row_idx, column=5, value=q.get("helper", ""))
    # Options: pretty-print als "value | label" per regel
    opts = q.get("options", [])
    opts_text = "\n".join(f"{o.get('value', '')} | {o.get('label', '')}" for o in opts)
    cell = ws.cell(row=row_idx, column=6, value=opts_text)
    cell.alignment = WRAP
    for col_idx in range(1, 8):
        ws.cell(row=row_idx, column=col_idx).border = BORDER
        ws.cell(row=row_idx, column=col_idx).alignment = WRAP
    ws.row_dimensions[row_idx].height = max(60, 20 + 15 * len(opts))

ws.freeze_panes = "A2"


# ========= Tabblad 3: Topics =========
ws = wb.create_sheet("Topics")
columns = [
    ("id",               "KEY",  7,  "Topic-ID, gebruikt in rules.activatedTopicIds"),
    ("subject",          "EDIT", 45, "De titel zoals in rapport"),
    ("category",         "EDIT", 30, "Categorie-tag (zichtbaar in rapport)"),
    ("type",             "EDIT", 30, "Type-tag (bv 'EU-verordening', 'Private certificering')"),
    ("esgPillar",        "EDIT", 10, "E / S / G / combinaties"),
    ("directIndirect",   "EDIT", 35, "Directe of indirecte impact; gebruikt voor hard-block"),
    ("description",      "EDIT", 60, "Korte toelichting (2-3 zinnen) in rapport"),
    ("whenRelevant",     "EDIT", 80, "Tweede zin van 'Waarom voor u relevant'"),
    ("triggerCriteria",  "EDIT", 35, "Intern hulpveld; niet zichtbaar in rapport"),
    ("sectorSpecific",   "EDIT", 10, "true/false"),
    ("sectors",          "EDIT", 35, "Sector-lijst"),
    ("priority",         "EDIT", 10, "Laag / Middel / Hoog"),
    ("horizon",          "EDIT", 22, "Nu / Binnen 1-3 jaar / Monitor / combinaties"),
    ("recommendedLabel", "EDIT", 35, "KRITIEK: Nu verplicht / Hoog relevant via keten / Marktstandaard / aanbevolen / Sectorspecifiek vervolgonderzoek / Monitoren / voorbereiden / Mogelijk relevant · nadere toetsing aanbevolen / Niet prioritair"),
    ("url",              "EDIT", 45, "Bronlink (primaire bron aanbevolen)"),
    ("note",             "EDIT", 45, "Interne notitie; niet zichtbaar in rapport"),
    ("Opmerking",        "NOTE", 60, "Typ hier je wijzigingsvoorstel of opmerking voor Act Right / Claude"),
]
for col_idx, (name, kind, width, descr) in enumerate(columns, start=1):
    apply_header(ws, col_idx, name, kind, width, descr)

for row_idx, t in enumerate(topics, start=2):
    ws.cell(row=row_idx, column=1,  value=t.get("id"))
    ws.cell(row=row_idx, column=2,  value=t.get("subject", ""))
    ws.cell(row=row_idx, column=3,  value=t.get("category", ""))
    ws.cell(row=row_idx, column=4,  value=t.get("type", ""))
    ws.cell(row=row_idx, column=5,  value=t.get("esgPillar", ""))
    ws.cell(row=row_idx, column=6,  value=t.get("directIndirect", ""))
    ws.cell(row=row_idx, column=7,  value=t.get("description", ""))
    ws.cell(row=row_idx, column=8,  value=t.get("whenRelevant", ""))
    ws.cell(row=row_idx, column=9,  value=t.get("triggerCriteria", ""))
    ws.cell(row=row_idx, column=10, value=str(t.get("sectorSpecific", False)).lower())
    ws.cell(row=row_idx, column=11, value=t.get("sectors", ""))
    ws.cell(row=row_idx, column=12, value=t.get("priority", ""))
    ws.cell(row=row_idx, column=13, value=t.get("horizon", ""))
    ws.cell(row=row_idx, column=14, value=t.get("recommendedLabel", ""))
    ws.cell(row=row_idx, column=15, value=t.get("url", ""))
    ws.cell(row=row_idx, column=16, value=t.get("note", ""))
    for col_idx in range(1, len(columns) + 1):
        ws.cell(row=row_idx, column=col_idx).border = BORDER
        ws.cell(row=row_idx, column=col_idx).alignment = WRAP
    # Dynamic row height based on longest content
    longest = max(len(str(t.get(k, ""))) for k, _, _, _ in columns)
    ws.row_dimensions[row_idx].height = min(200, max(40, longest // 3))

ws.freeze_panes = "C2"


# ========= Tabblad 4: Rules =========
ws = wb.create_sheet("Rules")
columns = [
    ("id",                 "KEY",  10, "Rule-ID"),
    ("questionId",         "KEY",  10, "Op welke vraag triggert deze rule"),
    ("trigger",            "KEY",  35, "Welk antwoord triggert"),
    ("themeId",            "KEY",  10, "Welk thema wordt versterkt (T1-T8)"),
    ("ruleType",           "EDIT", 20, "Versterk thema / Verzwak thema / etc."),
    ("scoreImpact",        "EDIT", 12, "Getal (meestal 1-4)"),
    ("activatedTopicIds",  "KEY",  28, "Lijst van topic-ID's die deze rule activeert"),
    ("preferredLabel",     "EDIT", 35, "KRITIEK: wat deze rule wil dat het label wordt"),
    ("preferredHorizon",   "EDIT", 22, "Wat deze rule wil dat de horizon wordt"),
    ("reportPhrasing",     "EDIT", 90, "Eerste zin van 'Waarom voor u relevant' — wordt gecombineerd met topic.whenRelevant"),
    ("Opmerking",          "NOTE", 60, "Typ hier je wijzigingsvoorstel of opmerking voor Act Right / Claude"),
]
for col_idx, (name, kind, width, descr) in enumerate(columns, start=1):
    apply_header(ws, col_idx, name, kind, width, descr)

for row_idx, r in enumerate(rules, start=2):
    ws.cell(row=row_idx, column=1,  value=r.get("id", ""))
    ws.cell(row=row_idx, column=2,  value=r.get("questionId", ""))
    ws.cell(row=row_idx, column=3,  value=r.get("trigger", ""))
    ws.cell(row=row_idx, column=4,  value=r.get("themeId", ""))
    ws.cell(row=row_idx, column=5,  value=r.get("ruleType", ""))
    ws.cell(row=row_idx, column=6,  value=r.get("scoreImpact"))
    ids = r.get("activatedTopicIds", [])
    ws.cell(row=row_idx, column=7,  value=", ".join(str(i) for i in ids) if ids else "")
    ws.cell(row=row_idx, column=8,  value=r.get("preferredLabel", ""))
    ws.cell(row=row_idx, column=9,  value=r.get("preferredHorizon", ""))
    ws.cell(row=row_idx, column=10, value=r.get("reportPhrasing", ""))
    for col_idx in range(1, len(columns) + 1):
        ws.cell(row=row_idx, column=col_idx).border = BORDER
        ws.cell(row=row_idx, column=col_idx).alignment = WRAP
    longest_phrasing = len(str(r.get("reportPhrasing", "")))
    ws.row_dimensions[row_idx].height = min(150, max(30, longest_phrasing // 2))

ws.freeze_panes = "D2"


# ========= Tabblad 5: Themas =========
ws = wb.create_sheet("Themas")
columns = [
    ("id",          "KEY",  8,  "Thema-ID (T1-T8)"),
    ("name",        "EDIT", 40, "Thema-naam (zichtbaar in rapport)"),
    ("description", "EDIT", 90, "Uitleg van het thema"),
    ("scope",       "EDIT", 45, "Scope-tag"),
    ("Opmerking",   "NOTE", 60, "Typ hier je wijzigingsvoorstel of opmerking voor Act Right / Claude"),
]
for col_idx, (name, kind, width, descr) in enumerate(columns, start=1):
    apply_header(ws, col_idx, name, kind, width, descr)

for row_idx, th in enumerate(themes, start=2):
    ws.cell(row=row_idx, column=1, value=th.get("id", ""))
    ws.cell(row=row_idx, column=2, value=th.get("name", ""))
    ws.cell(row=row_idx, column=3, value=th.get("description", ""))
    ws.cell(row=row_idx, column=4, value=th.get("scope", ""))
    for col_idx in range(1, 6):
        ws.cell(row=row_idx, column=col_idx).border = BORDER
        ws.cell(row=row_idx, column=col_idx).alignment = WRAP
    ws.row_dimensions[row_idx].height = max(40, len(str(th.get("description", ""))) // 2)

ws.freeze_panes = "B2"


# --- Opslaan ---
wb.save(OUTPUT)
print(f"\nKlaar: {OUTPUT}")
print(f"Bestand-grootte: {OUTPUT.stat().st_size // 1024} KB")
